from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from config import Parameters, ControlVars

# Expressions
@dataclass
class Configuation_expr:

    # Scalar
    confExString_TerminatingCondition: str
    confExString_InitialRateConfigurationNumber: str = "1"
    
    # Vector // we must init to empty // dict[label] -> expr_str
    confExString_InitialLevelValue: Dict[str, str] = field(default_factory=dict)
    confExString_InitialTimerValue: Dict[str, str] = field(default_factory=dict)
    confExString_InitialDiscretelyDynamicalNumericalVariableValue: Dict[str, str] = field(default_factory=dict)
    confExString_InitialCategoricalVariableValue: Dict[str, str] = field(default_factory=dict)

    # Matrix dict[row_label][col_label] -> expr_str
    confExString_LevelRate: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerLevelThreshold: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperLevelThreshold: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerLevelResultantRateConfiguration: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperLevelResultantRateConfiguration: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerLevelAssignmentAddress: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperLevelAssignmentAddress: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_TimerRate: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerTimerThreshold: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperTimerThreshold: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerTimerResultantRateConfiguration: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperTimerResultantRateConfiguration: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_LowerTimerAssignmentAddress: Dict[str, Dict[str, str]] = field(default_factory=dict)
    confExString_UpperTimerAssignmentAddress: Dict[str, Dict[str, str]] = field(default_factory=dict)
    
    confExString_AssignmentSequence: Dict[str, Dict[str, str]] = field(default_factory=dict)


#order matters since we do drs_timer(int)
#Now we want to be able to search here quickly. sooo maybe a dict with L/T/N/C as a first char.  similar to the last decide of the model
# may simplify some work later on 
@dataclass(frozen=True)
class DRSVars:

    # Levels
    level_labels: Tuple[str, ...] = (
        "OreExtraction_Level",
        "OreExtractedFromCurrentParcel_Level",
        "OreStock_Level",
        "Ore1Stock_Level",
        "Ore2Stock_Level",
    )

    # Timers
    timer_labels: Tuple[str, ...] = (
        "TimeExecutedInCurrentCampaignOrShutdown_Timer",
        "TimeExecutedInCurrentContingencySegment_Timer",
        "TimeInModeA_Timer",
        "TimeInModeAContingency_Timer",
        "TimeInModeAMineSurging_Timer",
        "TimeInModeB_Timer",
        "TimeInModeBContingency_Timer",
        "TimeInModeBMineSurging_Timer",
        "TimeInShutdown_Timer"
    )
    
    dynamic_numerical_labels: Tuple[str, ...] = (
        "MassOfCurrentParcel",
        "PercentageOfOre2InCurrentParcel",
        "NextParcelIsNewFacies"
    )

    # Since we had cat. in the decide of the last part of the model
    categorical_labels: Tuple[str, ...] = ("DummerCategorical",)

    def build_dicts(self) -> Dict[str, Tuple[str, int]]:
        # L : level, T : Timer, N : dynamic.., C : catego.

        m: Dict[str, Tuple[str, int]] = {}
        for i, s in enumerate(self.level_labels):
            m[s] = ("L", i)
        for i, s in enumerate(self.timer_labels):
            m[s] = ("T", i)
        for i, s in enumerate(self.dynamic_numerical_labels):
            m[s] = ("N", i)
        for i, s in enumerate(self.categorical_labels):
            m[s] = ("C", i)

        return m




# Variables
@dataclass
class Dimensions:
    dim_NumberOfLevels: int = 5
    dim_NumberOfTimerss: int = 9
    dim_NumberOfDiscretelyDynamicalNumericalVariabless: int = 3
    dim_NumberOfCategoricalVariabless: int = 1# was 1 on arena, should be 0/null will test all and see
    dim_NumberOfRateConfigurationss: int = 7
    dim_NumberOfAssignmentSequenceAddressess: int = 6
    dim_MaxLengthOfAssignmentSequences: int = 7


@dataclass
class DRSState:
    drs_RateConfigurationNumber: int = 1
    drs_TimeOfPreviousDRSUpdate: float = 0.0 # we start at 0
    drs_DurationUntilThresholdCrossing: float = 99999.0
    drs_ThresholdCrossingLevelOrTimerNumber: int = 0 #just init at 0 else it wont run, same for the otehr 2
    drs_ThresholdIsCrossedByTimer: int = 0
    drs_DirectionOfThresholdCrossing: int = 0
    

    # must have alocator to work
    drs_Level: List[float] = field(default_factory=list)
    drs_Timer: List[float] = field(default_factory=list)
    drs_DiscretelyDynamicalNumericalVariable: List[float] = field(default_factory=list)
    drs_CategoricalVariable: List[str] = field(default_factory=list)

    TNOW: float = 0.0

    def init(self, dim : Dimensions):
        self.drs_Level = [0.0] * dim.dim_NumberOfLevels
        self.drs_Timer = [0.0] * dim.dim_NumberOfTimerss
        self.drs_DiscretelyDynamicalNumericalVariable = [0.0] * dim.dim_NumberOfDiscretelyDynamicalNumericalVariabless
        self.drs_CategoricalVariable = [""] * dim.dim_NumberOfCategoricalVariabless


# Must write to get info from excel file
# Would use pandas howver pandas is not really suited when an excel has mustliple types, 
# using openyxl is more suited since we can do "type matching"

def to_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()

# input worksheet and output rows
def read_rows(ws :Worksheet) -> List[List[Any]]:
    rows : List[List[Any]] = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows

def find_row(rows : List[List[Any]], key: str) -> int :
    for i, row in enumerate(rows):
        if to_str(row[0]).__contains__(key): # given that some excel rows have name then extra info
            return i
    return -1

# May have to do testing for the parsing, think about the ""
def parse_scalar(rows: List[List[Any]], start_i: int) -> str:
    val = rows[start_i][2]
    s = to_str(val)
    return s

def parse_vector(rows: List[List[Any]], start_i: int) -> Dict[str, str]:
    # Col A empty
    # Col B label
    # Col C info

    # Will have us output a dict, with the labels, may decide later to just take the index
    # starting with label will make things easier to debug with the cost of running slower
    # TODO : change to reg arr with idx
    
    out : Dict[str, str] = {}
    i = start_i + 1

    while i < len(rows) :
        a = to_str(rows[i][0])
        b = to_str(rows[i][1]) if len(rows[i]) > 1 else ""
        c = to_str(rows[i][2]) if len(rows[i]) > 2 else ""

        if a != "":
            break
        if b=="" and c=="":
            break
        if b!="" and c!="":
            out[b]=c
        i += 1

    return out

def parse_matrix(rows: List[List[Any]], start_i: int) -> Dict[str, Dict[str, str]]:
    # Somewhat similar to vec
    # col B label
    # Row i + 1 cols
    # Rest fields
    # print(start_i)

    i = start_i + 1

    # Get headers
    col_labels = [to_str(x) for x in rows[i][2:] if to_str(x) != ""]

    i += 1

    out : Dict[str, Dict[str, str]] = {}

    while i < len(rows):
        a = to_str(rows[i][0])
        b = to_str(rows[i][1])
        c = to_str(rows[i][2])

        if a != "":
            break # we are done with 
        if b == "" or c == "":
            break # end of blocl
        if b != "":
            out[b] = {}
            for j, col in enumerate(col_labels):
                cell = rows[i][2+j] if (2+j) < len(rows[i]) else None
                val = to_str(cell)
                if val != "":
                    out[b][col] = val
        i += 1

    # print("keys:", list(out.keys())[:5])

    return out

def load_confExStrings_from_excel(xlsx_path: str, sheet_name: str | None = None) -> Configuation_expr:

    # Open excel sheet and read all rows.

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = read_rows(ws)

    index_terminatingCond = find_row(rows, "confExString_TerminatingCondition")
    index_initRateConfig = find_row(rows, "confExString_InitialRateConfigurationNumber")

    conf = Configuation_expr(confExString_TerminatingCondition="")
    conf.confExString_TerminatingCondition = parse_scalar(rows, index_terminatingCond)
    conf.confExString_InitialRateConfigurationNumber = parse_scalar(rows, index_initRateConfig)

    # Vectors

    vectors = [
        "confExString_InitialLevelValue",
        "confExString_InitialTimerValue",
        "confExString_InitialDiscretelyDynamicalNumericalVariableValue",
        "confExString_InitialCategoricalVariableValue"
    ]

    for vec in vectors:
        k_i = find_row(rows, vec)
        setattr(conf, vec, parse_vector(rows, k_i))


    matrices = [
        "confExString_LevelRate",
        "confExString_LowerLevelThreshold",
        "confExString_UpperLevelThreshold",
        "confExString_LowerLevelResultantRateConfiguration",
        "confExString_UpperLevelResultantRateConfiguration",
        "confExString_LowerLevelAssignmentAddress",
        "confExString_UpperLevelAssignmentAddress",
        "confExString_TimerRate",
        "confExString_LowerTimerThreshold",
        "confExString_UpperTimerThreshold",
        "confExString_LowerTimerResultantRateConfiguration",
        "confExString_UpperTimerResultantRateConfiguration",
        "confExString_LowerTimerAssignmentAddress",
        "confExString_UpperTimerAssignmentAddress",
        "confExString_AssignmentSequence"
    ]

    for matrix in matrices:
        k_i = find_row(rows, matrix)
        setattr(conf, matrix, parse_matrix(rows, k_i))

    return conf


    # Eeach row is the first index is used then we have a coef stored

def main():
    load_confExStrings_from_excel("MiningSystemDRS_ConfExStrings_v4.xlsx")

if __name__ == "__main__":
    main()